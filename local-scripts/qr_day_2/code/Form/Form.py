import os
import openpyxl
from datetime import datetime, timezone, timedelta
import firebase_admin
from firebase_admin import credentials, firestore

# --- 1. ฟังก์ชันช่วยสร้าง เวลามาตรฐานประเทศไทย (ISO 8601) ---
def get_thai_iso_string():
    tz_thai = timezone(timedelta(hours=7))
    now_thai = datetime.now(tz_thai)
    return now_thai.strftime('%Y-%m-%dT%H:%M:%S.%f')[:-3] + '+07:00'

# --- 2. ฟังก์ชันอ่านข้อมูลรหัสนักศึกษา ชื่อ และนามสกุลจากไฟล์ Excel (Col A, B, C) ---
def read_student_entries_from_excel(excel_path):
    print("=" * 65)
    print(f"📊 อ่านข้อมูลจากไฟล์ Excel: {os.path.basename(excel_path)}")
    print("=" * 65)

    if not os.path.exists(excel_path):
        print(f"❌ ไม่พบไฟล์ Excel ที่: {excel_path}")
        return []

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if 'StudentId' not in wb.sheetnames:
        print("❌ ไม่พบแท็บ 'StudentId' ในไฟล์ Excel")
        return []

    sheet = wb['StudentId']
    student_entries = []
    
    # อ่านเฉพาะ Column A, B, C ตั้งแต่แถวที่ 2 ถึงแถวสุดท้ายที่พบข้อมูล
    max_r = sheet.max_row
    for r in range(2, max_r + 1):
        cell_a = sheet.cell(row=r, column=1).value
        cell_b = sheet.cell(row=r, column=2).value
        cell_b_fname = str(cell_b).strip() if cell_b is not None else ''
        cell_c = sheet.cell(row=r, column=3).value
        cell_c_lname = str(cell_c).strip() if cell_c is not None else ''

        clean_sid = str(cell_a).split('.')[0].strip() if cell_a is not None else ''

        # หากมีข้อมูลใน Col A, B หรือ C
        if clean_sid or cell_b_fname or cell_c_lname:
            student_entries.append({
                'row': r,
                'studentId': clean_sid,
                'firstName': cell_b_fname,
                'lastName': cell_c_lname
            })

    print(f"✅ อ่านข้อมูลจากแท็บ StudentId (Col A, B, C แถวที่ 2 ถึง {max_r}) ได้ทั้งหมด {len(student_entries)} รายการ")
    print("=" * 65)
    return student_entries

# --- 3. ฟังก์ชันค้นหาข้อมูลนิสิตใน Firestore DB (Multi-Pass ID + Name Lookup) ---
def find_student_in_db(db, sid, first_name='', last_name=''):
    clean_sid = sid.strip() if isinstance(sid, str) else str(sid).strip()
    clean_fname = first_name.strip() if isinstance(first_name, str) else str(first_name).strip()
    clean_lname = last_name.strip() if isinstance(last_name, str) else str(last_name).strip()

    # 1. ค้นหาจาก studentId
    if clean_sid:
        try:
            query = db.collection('users').where('studentId', '==', clean_sid).limit(1).get()
            if query:
                doc = query[0]
                return {'id': doc.id, 'data': doc.to_dict(), 'match_method': 'Student ID'}
        except Exception:
            pass

        # 2. ค้นหาจาก line_uid
        try:
            query = db.collection('users').where('line_uid', '==', clean_sid).limit(1).get()
            if query:
                doc = query[0]
                return {'id': doc.id, 'data': doc.to_dict(), 'match_method': 'Line UID'}
        except Exception:
            pass

        # 3. ค้นหาโดยใช้ Document ID
        try:
            doc_ref = db.collection('users').document(clean_sid).get()
            if doc_ref.exists:
                return {'id': doc_ref.id, 'data': doc_ref.to_dict(), 'match_method': 'Document ID'}
        except Exception:
            pass

        # 4. ค้นหาจาก short_code / shortCode / walkin_temp_short_code / used_short_codes
        for field in ['short_code', 'shortCode', 'walkin_temp_short_code']:
            try:
                query = db.collection('users').where(field, '==', clean_sid).limit(1).get()
                if query:
                    doc = query[0]
                    return {'id': doc.id, 'data': doc.to_dict(), 'match_method': f'ShortCode ({field})'}
            except Exception:
                pass

        try:
            query = db.collection('users').where('used_short_codes', 'array_contains', clean_sid).limit(1).get()
            if query:
                doc = query[0]
                return {'id': doc.id, 'data': doc.to_dict(), 'match_method': 'Used ShortCodes'}
        except Exception:
            pass

    # 5. ค้นหาจาก ชื่อ และ นามสกุล (Col B & Col C)
    if clean_fname and clean_lname:
        try:
            query = db.collection('users').where('firstName', '==', clean_fname).where('lastName', '==', clean_lname).limit(1).get()
            if query:
                doc = query[0]
                return {'id': doc.id, 'data': doc.to_dict(), 'match_method': 'Exact Name (Col B & C)'}
        except Exception:
            pass

    # 6. ค้นหาจาก ชื่อ (Col B)
    if clean_fname:
        try:
            query = db.collection('users').where('firstName', '==', clean_fname).limit(1).get()
            if query:
                doc = query[0]
                return {'id': doc.id, 'data': doc.to_dict(), 'match_method': 'FirstName (Col B)'}
        except Exception:
            pass

    # 7. Fallback Name Mapping สำหรับชื่อสะกดต่างแบบ (ภาษาไทย/อังกฤษ)
    NAME_FALLBACK_MAP = {
        'Thae ThinZar': 'Ucfff2460334f1d377e59d9fb55322cbd',
        'Jaturon': 'Ud0d20c235ef57ec3320f2c871826a3ba',
        'Chollada': 'U0dc44887588f645704f53a5b2972a0d0',
        'Shawn': 'U4d4ac0a8d954936770483724e9261ffe',
        'ปรินาภรณ์': 'Uced045ad2a44a6b070e32b1d608062f9',
    }
    for key_name, doc_id in NAME_FALLBACK_MAP.items():
        if (clean_fname and key_name.lower() in clean_fname.lower()) or (clean_lname and key_name.lower() in clean_lname.lower()):
            try:
                doc_ref = db.collection('users').document(doc_id).get()
                if doc_ref.exists:
                    return {'id': doc_ref.id, 'data': doc_ref.to_dict(), 'match_method': 'Fuzzy Name Fallback'}
            except Exception:
                pass

    return None

# --- 4. ฟังก์ชันประมวลผลอนุมัติ Walk-in และบันทึก Day 2 Morning Check-in ---
def process_form_day2_checkin(db, student_entries, dry_run=True):
    print("\n" + "=" * 65)
    print("🚀 เริ่มประมวลผล Walk-in Approval & Day 2 Morning Check-in")
    print("=" * 65)

    walkin_approved_count = 0
    walkin_already_approved_count = 0

    checkin_success_count = 0
    checkin_already_checked_in_count = 0
    not_found_list = []

    total_entries = len(student_entries)

    if total_entries == 0:
        print("ไม่มีรายการรหัสนักศึกษาให้ประมวลผล")
        return {
            'walkin_approved': 0,
            'walkin_already_approved': 0,
            'checkin_success': 0,
            'checkin_already_checked_in': 0,
            'not_found_list': [],
            'total': 0,
            'is_100_percent': False
        }

    for item in student_entries:
        row_num = item['row']
        sid = item['studentId']
        fname = item['firstName']
        lname = item['lastName']

        student = find_student_in_db(db, sid, fname, lname)

        if not student:
            print(f" 🔴 [แถว {row_num}] หาผู้ใช้ไม่พบใน DB: รหัส {sid} | ชื่อ-สกุล: {fname} {lname}")
            not_found_list.append({
                'row': row_num,
                'studentId': sid,
                'firstName': fname,
                'lastName': lname
            })
            continue

        student_id = student['id']
        student_data = student['data']
        match_method = student.get('match_method', 'Auto')
        db_first_name = student_data.get('firstName', '')
        db_last_name = student_data.get('lastName', '')
        std_id_code = student_data.get('studentId', sid)
        full_name = f"{db_first_name} {db_last_name}".strip() or f"{fname} {lname}".strip() or "ไม่ระบุชื่อ"
        
        note = str(student_data.get('note', ''))
        walkin_temp_code = student_data.get('walkin_temp_short_code')
        walkin_status = student_data.get('walkin_status', '')
        walkin_verified = student_data.get('walkin_verified', False)

        # -------------------------------------------------------------
        # STEP 1: ตรวจสอบและอนุมัติ Walk-in หากมี Note "รอบหน้างาน"
        # -------------------------------------------------------------
        if 'รอบหน้างาน' in note or walkin_temp_code or walkin_status or walkin_verified:
            if walkin_status == 'APPROVED' or walkin_verified is True:
                walkin_already_approved_count += 1
            else:
                # ทำการ อนุมัติ Walk-in ก่อน
                iso_timestamp = get_thai_iso_string()
                approval_payload = {
                    'walkin_status': 'APPROVED',
                    'walkin_verified': True,
                    'walkin_approved_at': iso_timestamp,
                    'walkin_approved_by_staff_name': 'Admin Script CLI',
                    'walkin_approved_by_staff_uid': 'ADMIN_SCRIPT',
                    'updatedAt': iso_timestamp
                }

                staff_log_payload = {
                    'action': 'APPROVE_WALKIN',
                    'actor_name': 'Admin Script CLI',
                    'actor_uid': 'ADMIN_SCRIPT',
                    'actor_username': 'admin_cli',
                    'target_name': full_name,
                    'target_student_id': std_id_code,
                    'target_user_id': student_id,
                    'timestamp': iso_timestamp,
                    'details': 'อนุมัติผู้สมัคร Walk-in (รอบหน้างาน) จากแบบฟอร์มตอบกลับ'
                }

                if not dry_run:
                    user_ref = db.collection('users').document(student_id)
                    user_ref.update(approval_payload)

                    log_ref = db.collection('staff_access_logs').document()
                    log_ref.set(staff_log_payload)

                    print(f" 🚶 [บันทึกจริง] อนุมัติ Walk-in (รอบหน้างาน) สำเร็จ: {full_name} ({std_id_code})")
                else:
                    print(f" 🚶 [Dry-Run] พร้อมอนุมัติ Walk-in (รอบหน้างาน): {full_name} ({std_id_code})")

                walkin_approved_count += 1

        # -------------------------------------------------------------
        # STEP 2: บันทึก Day 2 Morning Check-in
        # -------------------------------------------------------------
        if student_data.get('checkin_day2_morning'):
            print(f" 🟡 [แถว {row_num}] มี Log แล้ว ไม่ต้องบันทึกเพิ่ม: {full_name} ({std_id_code}) [{match_method}]")
            checkin_already_checked_in_count += 1
            continue

        iso_timestamp = get_thai_iso_string()
        user_update_payload = {
            'checkin_day2_morning': iso_timestamp,
            'checkin_day2_morning_by': 'Admin Script CLI',
            'checkin_day2_morning_by_staff_uid': 'ADMIN_SCRIPT',
            'checkin_day2_morning_by_staff_pic': '',
            'checkin_day2_morning_by_staff_username': 'admin_cli',
            'checkin_day2_morning_operator_user': 'admin_cli',
            'checkin_day2_morning_search_method': 'FORM_EXCEL',
            'checkin_day2_morning_ip': '127.0.0.1',
            'checkin_day2_morning_device_model': 'Python Excel Form Script',
            'checkin_day2_morning_user_agent': 'Python-openpyxl',
            'checkin_day2_morning_platform': 'macOS Python',
            'updatedAt': iso_timestamp
        }

        short_code_val = student_data.get('short_code') or student_data.get('walkin_temp_short_code') or ''

        log_payload = {
            'checkin_by_staff_name': 'Admin Script CLI',
            'checkin_by_staff_uid': 'ADMIN_SCRIPT',
            'checkin_by_staff_username': 'admin_cli',
            'checkin_type': 'DAY2_MORNING',
            'device_model': 'Python Excel Form Script',
            'department': student_data.get('department', ''),
            'firstName': db_first_name or fname,
            'lastName': db_last_name or lname,
            'studentId': std_id_code,
            'short_code': short_code_val,
            'ip': '127.0.0.1',
            'operator_user': 'admin_cli',
            'platform': 'macOS Python',
            'search_method': 'FORM_EXCEL',
            'timestamp': iso_timestamp,
            'user_agent': 'Python-openpyxl',
            'user_doc_id': student_id
        }

        if not dry_run:
            user_ref = db.collection('users').document(student_id)
            user_ref.update(user_update_payload)

            log_ref = db.collection('registration_checkin_logs').document()
            log_ref.set(log_payload)

            print(f" 🟢 [บันทึกจริง] บันทึกเช็คอิน Day 2 เช้าสำเร็จ: {full_name} ({std_id_code}) [{match_method}]")
        else:
            print(f" 🟢 [Dry-Run] พร้อมบันทึกเช็คอิน Day 2 เช้า: {full_name} ({std_id_code}) [{match_method}]")

        checkin_success_count += 1

    not_found_count = len(not_found_list)

    print("-" * 65)
    print("📈 สรุปผลการประมวลผลการตอบแบบฟอร์ม Day 2")
    print(f" 🚶 การอนุมัติสิทธิ์ Walk-in (รอบหน้างาน):")
    print(f"   • อนุมัติใหม่สำเร็จ: {walkin_approved_count} คน")
    print(f"   • มี Log อนุมัติอยู่แล้ว: {walkin_already_approved_count} คน")
    print(f" 🌅 การลงทะเบียน Day 2 Morning Check-in:")
    print(f"   🟢 1. ลงทะเบียนใหม่สำเร็จ: {checkin_success_count} คน")
    print(f"   🟡 2. มี Log อยู่แล้ว (ไม่ต้องบันทึกเพิ่ม): {checkin_already_checked_in_count} คน")
    print(f"   🔴 3. หาผู้ใช้ไม่พบใน DB: {not_found_count} คน")
    print("-" * 65)
    print(f" 📦 รวมประมวลผลทั้งหมด: {total_entries} / {total_entries} รายการ")

    is_100_percent = (checkin_success_count + checkin_already_checked_in_count == total_entries) and (total_entries > 0)
    if is_100_percent:
        print(" ✅ ตรวจสอบความถูกต้อง: ประมวลผลครบถ้วนทุกรายการ (100%)")
    else:
        print(f" ⚠️ มีบางรายการที่หาใน DB ไม่พบ ({not_found_count} รายการ)")

    print("=" * 65)

    return {
        'walkin_approved': walkin_approved_count,
        'walkin_already_approved': walkin_already_approved_count,
        'checkin_success': checkin_success_count,
        'checkin_already_checked_in': checkin_already_checked_in_count,
        'not_found_list': not_found_list,
        'total': total_entries,
        'is_100_percent': is_100_percent
    }

# --- 5. ส่วนการทำงานหลัก (Main Execution) ---
if __name__ == '__main__':
    KEY_PATH = '/Users/buaboocha.bs/Documents/Rak-File/Web_smo/Dev/frontend-reg/local-scripts/smo-vidva-bangmod-firebase-adminsdk-fbsvc-3543e8d9ee.json'
    EXCEL_PATH = '/Users/buaboocha.bs/Documents/Rak-File/Web_smo/Dev/frontend-reg/local-scripts/qr_day_2/แบบทดสอบท้ายกิจกรรม (Engineering Post-Orientation Form)  (การตอบกลับ).xlsx'

    print("🔍 ตรวจสอบพาธไฟล์ Excel:", os.path.abspath(EXCEL_PATH))

    student_entries = read_student_entries_from_excel(EXCEL_PATH)

    if student_entries:
        if not os.path.exists(KEY_PATH):
            print(f"❌ ไม่พบไฟล์ Firebase Key ที่: {KEY_PATH}")
        else:
            if not firebase_admin._apps:
                cred = credentials.Certificate(KEY_PATH)
                firebase_admin.initialize_app(cred)

            db = firestore.client()

            DRY_RUN = False  # กำหนด True เพื่อทดสอบอย่างเดียว หรือ False เพื่ออัพเดตลง Firestore จริง

            # -------------------------------------------------------------
            # เริ่มกระบวนการประมวลผล Walk-in Approval & Day 2 Morning Check-in
            # -------------------------------------------------------------
            summary = process_form_day2_checkin(db, student_entries, dry_run=DRY_RUN)

            # -------------------------------------------------------------
            # พิมพ์รายงานสรุปภาพรวมทั้งหมด (Grand Summary Report)
            # -------------------------------------------------------------
            not_found_list = summary['not_found_list']

            print("\n🏆=====================================================")
            print("📋 สรุปรายงานภาพรวมการประมวลผลทั้งหมด Day 2 จากแบบฟอร์ม (Grand Summary Report)")
            print("==================================================")
            print(f"📁 จำนวนรายการที่อ่านได้จาก Excel: {summary['total']} รายการ")
            print("--------------------------------------------------")
            print("🚶 1. การอนุมัติสิทธิ์ Walk-in (รอบหน้างาน):")
            print(f"  • อนุมัติใหม่สำเร็จ: {summary['walkin_approved']} คน")
            print(f"  • มี Log อนุมัติอยู่แล้ว: {summary['walkin_already_approved']} คน")
            print("--------------------------------------------------")
            print("🌅 2. การลงทะเบียน Day 2 Morning Check-in:")
            print(f"  • บันทึกใหม่สำเร็จ: {summary['checkin_success']} คน")
            print(f"  • มี Log ลงทะเบียนอยู่แล้ว: {summary['checkin_already_checked_in']} คน")
            print(f"  • หาไม่พบใน DB: {len(not_found_list)} คน")
            print("--------------------------------------------------")

            if not_found_list:
                print(f"❌ รายชื่อที่ไม่พบใน DB ทั้งหมด ({len(not_found_list)} รายการ):")
                for nf in not_found_list:
                    print(f"  • แถว {nf['row']}: รหัส {nf['studentId']} | ชื่อ-สกุล: {nf['firstName']} {nf['lastName']}")
                print("--------------------------------------------------")

            if summary['is_100_percent']:
                print("✅ สถานะภาพรวม: การประมวลผลสมบูรณ์ 100% ครบถ้วนทุกรายการ!")
            else:
                processed_total = summary['checkin_success'] + summary['checkin_already_checked_in']
                print(f"⚠️ สถานะภาพรวม: ประมวลผลได้ {processed_total} / {summary['total']} รายการ")
            print("=======================================================\n")
